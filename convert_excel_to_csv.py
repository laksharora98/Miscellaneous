"""
Convert the multi-sheet Excel trade list into a single clean CSV file.

Input:  TRADE LIST SONOSUMMIT 2025 FINAL 1.xlsx (6 sheets)
Output: trade_list_sonosummit_2025.csv

Each sheet shares 8 core columns:
  S.No, Company Name, Name, Designation, Email ID, Mobile No., IRIA Co ordinator, Address

The script adds a "Category" column derived from the sheet name, forward-fills
Company Name for grouped rows, and skips empty/title/header rows.
"""

import csv
import openpyxl

INPUT_FILE = "TRADE LIST SONOSUMMIT 2025 FINAL 1.xlsx"
OUTPUT_FILE = "trade_list_sonosummit_2025.csv"

HEADERS = [
    "Category",
    "S.No",
    "Company Name",
    "Name",
    "Designation",
    "Email ID",
    "Mobile No.",
    "IRIA Co ordinator",
    "Address",
]

SHEET_CATEGORY_MAP = {
    "MULTINATIONAL - TO SEND": "Multinational",
    "USG - TO SEND": "Ultrasound (USG)",
    "NON VASC TO SEND": "Non Vascular Intervention",
    "CONTRAST - TO BE SENT": "Contrast & Injectors",
    "INDIAN - TO BE SENT": "Indian",
    "MISCELLANEOUS- BOOKS SEND": "Miscellaneous & Books",
}


def clean_value(val):
    """Strip whitespace and control characters from strings; return empty string for None."""
    if val is None:
        return ""
    if isinstance(val, str):
        # Remove Unicode control characters (LTR/RTL marks, etc.)
        val = val.replace("\u200e", "").replace("\u200f", "")
        # Normalize newlines to semicolons (multi-value fields)
        val = val.replace("\r\n", "; ").replace("\n", "; ").replace("\r", "; ")
        val = val.strip()
        return val if val else ""
    return val


def format_mobile(val):
    """Convert mobile numbers to clean strings, avoiding scientific notation."""
    if val is None or val == "":
        return ""
    if isinstance(val, float):
        # Avoid scientific notation: 9.84e+09 -> "9840000000"
        if val == int(val):
            return str(int(val))
        return str(val)
    if isinstance(val, int):
        return str(val)
    # Already a string
    return str(val).strip()


def is_empty_row(values):
    """Check if all 8 core column values are empty/None."""
    return all(v is None or (isinstance(v, str) and v.strip() == "") for v in values)


def get_subcategory_rows(ws):
    """Detect sub-category rows via wide merged cell ranges (B through H or beyond).
    These are label rows like "MULTINATIONAL MULTI MODALITY COMPANIES" or
    "CONTRAST AGENTS" that span many columns but aren't actual contact entries.
    """
    subcategory_rows = set()
    for mc in ws.merged_cells.ranges:
        # Wide merges starting at/near column B, spanning 5+ columns, in data area (row >= 3)
        if mc.min_col <= 2 and (mc.max_col - mc.min_col) >= 4 and mc.min_row >= 3:
            subcategory_rows.add(mc.min_row)
    return subcategory_rows


def main():
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
    all_rows = []

    for sheet_name in wb.sheetnames:
        category = SHEET_CATEGORY_MAP.get(sheet_name, sheet_name)
        ws = wb[sheet_name]
        subcategory_rows = get_subcategory_rows(ws)

        last_company = ""

        for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_col=8, values_only=True), start=3):
            values = list(row)

            # Pad to 8 columns if sheet has fewer
            while len(values) < 8:
                values.append(None)

            if is_empty_row(values):
                continue

            if row_idx in subcategory_rows:
                continue

            # Clean all values
            values = [clean_value(v) for v in values]

            # Forward-fill Company Name
            if values[1]:
                last_company = values[1]
            else:
                values[1] = last_company

            # Format mobile number
            values[5] = format_mobile(values[5])

            # Build output row: Category + 8 core columns
            out_row = [category] + values
            all_rows.append(out_row)

    # Write CSV
    with open(OUTPUT_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        writer.writerows(all_rows)

    print(f"Wrote {len(all_rows)} data rows to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
